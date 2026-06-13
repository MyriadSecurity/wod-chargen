#!/usr/bin/env python3
"""Import MES character sheet lookup tables into wod_chargen JSON data files.

Reads the hidden ``tables`` and ``discipline_tables`` worksheets from the
Modern Enigma Society Laws of the Night spreadsheet (e.g. Monroe.xlsx).

Usage:
    uv run --with openpyxl python scripts/import_mes_tables.py [path/to/Monroe.xlsx]
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "wod_chargen" / "games" / "lotn_v5" / "data"
DEFAULT_XLSX = Path("/var/home/gscott/Downloads/Monroe.xlsx")
FALLBACK_XLSX = Path.home() / "Downloads" / "Monroe.xlsx"

PREDATOR_IDS = {
    "Alleycat",
    "Bagger",
    "Cleaver",
    "Consentualist",
    "Extortionist",
    "Farmer",
    "Ferryman",
    "Graverobber",
    "Hitcher",
    "Osiris",
    "Sandman",
    "Scene Queen",
    "Siren",
}


def slug(text: str) -> str:
    text = text.replace("´", "'").lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def _num(value: Any) -> int | float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value == int(value):
            return int(value)
        return value
    return None


def _row_list(row: tuple[Any, ...] | None, width: int = 12) -> list[Any]:
    values = list(row or [])
    while len(values) < width:
        values.append(None)
    return values


def resolve_xlsx(path: Path | None) -> Path:
    if path and path.exists():
        return path
    for candidate in (DEFAULT_XLSX, FALLBACK_XLSX):
        if candidate.exists():
            return candidate
    raise FileNotFoundError("Monroe.xlsx not found; pass path as first argument")


def sheet_version(wb: Any) -> str | None:
    try:
        ws = wb["start-here"]
        for coord in ("C2", "C1", "B2"):
            value = ws[coord].value
            if value and "MES" in str(value):
                return str(value)
    except (KeyError, AttributeError):
        pass
    return None


def import_clans(ws: Any) -> list[dict[str, Any]]:
    clans: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, max_row=18, values_only=True):
        row = _row_list(row, 12)
        name = row[1]
        if not name or name in ("Clans", "Generation"):
            continue
        bane_parts = [str(row[i]).strip() for i in range(9, 12) if row[i]]
        clans.append(
            {
                "id": slug(str(name)),
                "label": str(name),
                "enabled": row[0] if isinstance(row[0], bool) else None,
                "nickname": row[2],
                "disciplines": [d for d in (row[3], row[4], row[5]) if d],
                "archetypes": [a for a in (row[6], row[7], row[8]) if a],
                "bane": " ".join(bane_parts) if bane_parts else None,
                "compulsion": None,
            }
        )
    return clans


def import_generation(ws: Any) -> dict[str, Any]:
    generations: list[dict[str, Any]] = []
    blood_potency: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=20, max_row=40, values_only=True):
        row = _row_list(row, 8)
        label = row[1]
        if not label:
            continue
        if "Generation" in str(label) and row[2] is not None and "Blood Potency" not in str(label):
            if str(label) == "Generation":
                continue
            generations.append(
                {
                    "label": str(label),
                    "min_blood_potency": _num(row[2]),
                    "max_blood_potency": _num(row[3]),
                    "notes": row[4],
                    "generation_number": _num(row[5]),
                    "thin_blood": row[0] is True,
                    "ghoul": str(label).strip().lower() == "ghoul",
                }
            )
        if str(label) == "Blood Potency":
            continue
        rating = _num(row[1])
        if rating is not None and row[2] is not None:
            blood_potency.append(
                {
                    "rating": rating,
                    "blood_surge_bonus": _num(row[2]),
                    "mending_per_rouse": _num(row[3]),
                    "discipline_defense_bonus": _num(row[4]),
                    "discipline_rouse_bonus": row[5],
                }
            )
    return {"generations": generations, "blood_potency": blood_potency}


def import_predators(ws: Any) -> list[dict[str, Any]]:
    predators: list[dict[str, Any]] = []
    by_id: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=42, max_row=124, values_only=True):
        row = _row_list(row, 10)
        name = row[1]
        if name not in PREDATOR_IDS:
            continue
        if name not in by_id:
            entry = {
                "id": slug(name),
                "label": name,
                "humanity_mod": _num(row[2]) if row[2] is not None else None,
                "hunting_pool": None,
                "max_blood_potency": _num(row[5]),
                "grants": [],
            }
            by_id[name] = entry
            predators.append(entry)
        entry = by_id[name]
        if row[3] in ("Wits", "Manipulation", "Composure", "Charisma", "Strength", "Dexterity", "Intelligence"):
            entry["hunting_pool"] = {"attribute": slug(str(row[3])), "skill": slug(str(row[4]))}
            continue
        grant_type = row[2]
        if grant_type in ("Mortal Connections", "Merits", "Flaws"):
            grant: dict[str, Any] = {
                "kind": slug(grant_type),
                "category": row[3],
                "name": row[4],
                "dots": _num(row[5]),
            }
            if row[6]:
                grant["modifier"] = str(row[6])
            if row[7] is not None:
                grant["modifier_dots"] = _num(row[7])
            entry["grants"].append(grant)
    return predators


def import_dyscrasias(ws: Any) -> list[dict[str, Any]]:
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in ws.iter_rows(min_row=57, max_row=84, values_only=True):
        row = _row_list(row, 6)
        resonance, name, effect = row[1], row[2], row[3]
        if resonance not in ("Choleric", "Melancholic", "Phlegmatic", "Sanguine", "Animal", "None"):
            continue
        if not name or name == "None" or not effect:
            continue
        groups[str(resonance)].append({"name": str(name), "effect": str(effect).strip()})
    return [{"resonance": k, "dyscrasias": v} for k, v in groups.items()]


def import_background_catalog(ws: Any) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=231, max_row=275, values_only=True):
        row = _row_list(row, 8)
        if row[1] != "Mortal Connections" or not row[2]:
            continue
        modifier = row[3]
        cost = _num(row[4])
        max_dots = _num(row[5])
        kind = "connection"
        if modifier:
            kind = "disadvantage" if cost is not None and cost < 0 else "advantage"
        entries.append(
            {
                "background": slug(str(row[2])),
                "background_label": str(row[2]),
                "kind": kind,
                "modifier": str(modifier) if modifier else None,
                "modifier_id": slug(str(modifier)) if modifier else None,
                "dot_cost": cost,
                "max_dots": max_dots,
                "flag": row[6],
            }
        )
    return entries


def import_merits_flaws(ws: Any) -> dict[str, list[dict[str, Any]]]:
    merits: list[dict[str, Any]] = []
    flaws: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=349, max_row=427, values_only=True):
        row = _row_list(row, 8)
        kind = row[1]
        if kind not in ("Merits", "Flaws"):
            continue
        category, name = row[2], row[3]
        if not category or not name:
            continue
        entry = {
            "id": slug(str(name)),
            "label": str(name),
            "category": slug(str(category)),
            "category_label": str(category),
            "dot_cost": _num(row[4]),
            "max_dots": _num(row[5]),
            "thin_blood_only": row[0] is True,
            "ghoul_only": row[0] is False and str(category) == "Ghoul",
        }
        if kind == "Merits":
            merits.append(entry)
        else:
            flaws.append(entry)
    return {"merits": merits, "flaws": flaws}


def import_loresheets(ws: Any) -> list[dict[str, Any]]:
    sheets: dict[str, dict[str, Any]] = {}
    order: list[str] = []
    for row in ws.iter_rows(min_row=276, max_row=347, values_only=True):
        row = _row_list(row, 10)
        if row[1] != "Loresheet" or not row[2]:
            continue
        ls_name = str(row[2])
        if ls_name not in sheets:
            sheets[ls_name] = {
                "id": slug(ls_name),
                "label": ls_name,
                "restriction": None,
                "levels": [],
            }
            order.append(ls_name)
        sheet = sheets[ls_name]
        if row[7]:
            sheet["restriction"] = str(row[7])
        sheet["levels"].append(
            {
                "dots": int(_num(row[4]) or 0),
                "label": str(row[3]),
                "id": slug(str(row[3])),
            }
        )
    return [sheets[name] for name in order]


def import_xp_matrix(ws: Any) -> dict[str, Any]:
    header: list[Any] = []
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=428, max_row=439, values_only=True):
        row = _row_list(row, 10)
        if row[1] == "XP Table" or (row[2] is None and row[1] is None and not header):
            continue
        if not header and row[2] is not None:
            header = [_num(v) for v in row[2:8]]
            continue
        if row[1] is not None:
            rows.append({"new_level": _num(row[1]), "costs": [_num(v) for v in row[2:8]]})
    columns = ["background_merit", "skill", "attribute", "discipline_in", "discipline_out", "blood_potency"]
    return {"columns": columns, "column_keys": header, "rows": rows}


def import_equipment(ws: Any) -> list[dict[str, Any]]:
    qualities: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=443, max_row=470, values_only=True):
        row = _row_list(row, 8)
        category, name = row[1], row[2]
        if category not in ("Melee", "Ranged", "Protective", "Miscellaneous"):
            continue
        if not name:
            continue
        qualities.append(
            {
                "category": slug(str(category)),
                "category_label": str(category),
                "id": slug(str(name)),
                "label": str(name),
                "effect": row[3],
                "cost": _num(row[4]),
            }
        )
    return qualities


def import_discipline_tables(ws: Any) -> dict[str, Any]:
    meta: dict[str, dict[str, Any]] = {}
    powers: dict[str, list[dict[str, Any]]] = defaultdict(list)
    quick_ref: list[dict[str, Any]] = []

    for row in ws.iter_rows(values_only=True):
        row = _row_list(row, 36)
        disc_a = row[0]
        if disc_a and isinstance(disc_a, str) and disc_a not in ("Discipline",) and row[4] in (
            "Social",
            "Mental",
            "Physical",
            "Variable",
        ):
            meta[disc_a] = {
                "id": slug(disc_a),
                "label": disc_a,
                "in_clan": row[3],
                "type": row[4],
                "threat": row[5],
                "resonance": row[6],
            }

        pdisc, level, power = row[10], row[11], row[12]
        if pdisc and power and level is not None:
            entry = {
                "level": int(level) if isinstance(level, float) and level == int(level) else level,
                "id": slug(str(power)),
                "label": str(power),
                "amalgam": row[13],
                "amalgam_level": _num(row[14]),
                "prerequisite": row[15],
                "cost": row[16],
                "pool_attribute": row[17],
                "pool_skill": row[18],
                "pool_bonus": _num(row[19]),
                "difficulty": row[20],
                "duration": row[21],
                "notes": row[24],
            }
            powers[str(pdisc)].append(entry)

        if row[28] and row[29] and row[28] in (
            "Celerity",
            "Potence",
            "Presence",
            "Thin-Blood Alchemy",
            "Animalism",
            "Auspex",
            "Dominate",
            "Fortitude",
            "Obfuscate",
            "Oblivion",
            "Protean",
            "Blood Sorcery",
        ):
            quick_ref.append(
                {
                    "discipline": slug(str(row[28])),
                    "discipline_label": str(row[28]),
                    "power": str(row[29]),
                    "power_id": slug(str(row[29])),
                    "level": _num(row[26]),
                    "rouse": _num(row[30]),
                    "duration": row[35],
                }
            )

    disciplines: list[dict[str, Any]] = []
    for label in sorted(powers.keys()):
        disc_id = slug(label)
        named = [p for p in powers[label] if not p["label"].startswith("Counterfeit")]
        counterfeit = [p for p in powers[label] if p["label"].startswith("Counterfeit")]
        disciplines.append(
            {
                **meta.get(label, {"id": disc_id, "label": label}),
                "powers": powers[label],
                "power_count": len(powers[label]),
                "named_formula_count": len(named),
                "counterfeit_count": len(counterfeit),
            }
        )

    return {
        "disciplines": disciplines,
        "quick_reference": quick_ref,
    }


def write_json(path: Path, payload: dict[str, Any] | list[Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def _merge_merit_flaw_descriptions(merits_flaws: dict[str, list[dict[str, Any]]]) -> None:
    """Attach pocket-book descriptions and rules; preserve any already on disk."""
    from merit_flaw_rules import RULES
    from merits_flaws_descriptions import DESCRIPTIONS

    existing_desc: dict[str, str] = {}
    existing_rules: dict[str, dict[str, Any]] = {}
    path = DATA / "merits_flaws.json"
    if path.exists():
        prior = json.loads(path.read_text(encoding="utf-8"))
        for kind in ("merits", "flaws"):
            for entry in prior.get(kind, []):
                if entry.get("description"):
                    existing_desc[entry["id"]] = entry["description"]
                if entry.get("rules"):
                    existing_rules[entry["id"]] = entry["rules"]

    for kind in ("merits", "flaws"):
        for entry in merits_flaws.get(kind, []):
            entry_id = entry["id"]
            entry["description"] = (
                existing_desc.get(entry_id)
                or DESCRIPTIONS.get(entry_id)
                or entry.get("description")
            )
            rules = existing_rules.get(entry_id) or RULES.get(entry_id)
            if rules:
                entry["rules"] = rules
            else:
                entry.pop("rules", None)


def _merge_discipline_power_enrichment(discipline_data: dict[str, Any]) -> None:
    """Preserve enriched descriptions/rules; re-apply from scripts after MES import."""
    from enrich_discipline_powers import enrich

    existing_desc: dict[str, str] = {}
    path = DATA / "discipline_powers.json"
    if path.exists():
        prior = json.loads(path.read_text(encoding="utf-8"))
        for disc in prior.get("disciplines", []):
            for entry in disc.get("powers", []):
                if entry.get("description"):
                    existing_desc[entry["id"]] = entry["description"]

    for disc in discipline_data.get("disciplines", []):
        for entry in disc.get("powers", []):
            pid = entry["id"]
            if existing_desc.get(pid):
                entry["description"] = existing_desc[pid]

    enrich(discipline_data)


def sync_legacy_merits(merits_flaws: dict[str, list[dict[str, Any]]], source_ref: dict[str, Any]) -> None:
    """Keep merits.json compatible with generator until it loads merits_flaws.json."""
    write_json(
        DATA / "merits.json",
        {
            "source_ref": source_ref,
            "merits": [m["id"] for m in merits_flaws["merits"] if not m.get("thin_blood_only")],
            "flaws": [f["id"] for f in merits_flaws["flaws"] if not f.get("thin_blood_only")],
        },
    )


def sync_thin_blood_merits(merits_flaws: dict[str, list[dict[str, Any]]], source_ref: dict[str, Any]) -> None:
    thin_merits = [m for m in merits_flaws["merits"] if m.get("thin_blood_only")]
    thin_flaws = [f for f in merits_flaws["flaws"] if f.get("thin_blood_only")]
    write_json(
        DATA / "thin_blood_merits.json",
        {
            "source_ref": source_ref,
            "merits": [{"id": m["id"], "label": m["label"]} for m in thin_merits],
            "flaws": [{"id": f["id"], "label": f["label"]} for f in thin_flaws],
        },
    )


def sync_thin_blood_formulas(discipline_data: dict[str, Any], source_ref: dict[str, Any]) -> None:
    tba = next((d for d in discipline_data["disciplines"] if d.get("id") == "thin_blood_alchemy"), None)
    formulas: list[dict[str, Any]] = []
    if tba:
        for power in tba["powers"]:
            if power["label"].startswith("Counterfeit"):
                continue
            formulas.append(
                {
                    "id": power["id"],
                    "label": power["label"],
                    "level": power["level"],
                }
            )
    write_json(
        DATA / "thin_blood_formulas.json",
        {"source_ref": source_ref, "formulas": formulas},
    )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsx", nargs="?", type=Path, help="Path to Monroe.xlsx")
    parser.add_argument("--dry-run", action="store_true", help="Parse only; do not write files")
    args = parser.parse_args(argv)

    try:
        import openpyxl
    except ImportError:
        print("openpyxl required: uv run --with openpyxl python scripts/import_mes_tables.py", file=sys.stderr)
        return 1

    xlsx = resolve_xlsx(args.xlsx)
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    version = sheet_version(wb)
    imported_at = datetime.now(UTC).isoformat()
    source_ref = {
        "mes_sheet_version": version,
        "mes_workbook": xlsx.name,
        "imported_at": imported_at,
        "import_script": "scripts/import_mes_tables.py",
        "srd": "https://www.oneworldofdarkness.com/laws-of-the-night/",
    }

    tables = wb["tables"]
    clans = import_clans(tables)
    generation = import_generation(tables)
    predators = import_predators(tables)
    dyscrasias = import_dyscrasias(tables)
    background_catalog = import_background_catalog(tables)
    merits_flaws = import_merits_flaws(tables)
    _merge_merit_flaw_descriptions(merits_flaws)
    merits_flaws["source_ref"] = source_ref
    loresheets = {"source_ref": source_ref, "loresheets": import_loresheets(tables)}
    xp_matrix = {"source_ref": source_ref, **import_xp_matrix(tables)}
    equipment = {"source_ref": source_ref, "qualities": import_equipment(tables)}
    predator_catalog = {"source_ref": source_ref, "predators": predators}
    clans_table = {"source_ref": source_ref, "clans": clans}
    generation_data = {"source_ref": source_ref, **generation}
    dyscrasia_data = {"source_ref": source_ref, "resonances": import_dyscrasias(tables)}

    discipline_data = import_discipline_tables(wb["discipline_tables"])
    discipline_data["source_ref"] = source_ref
    _merge_discipline_power_enrichment(discipline_data)

    background_data = {"source_ref": source_ref, "entries": background_catalog}

    manifest = {
        "source_ref": source_ref,
        "files": {
            "merits_flaws.json": {
                "merits": len(merits_flaws["merits"]),
                "flaws": len(merits_flaws["flaws"]),
            },
            "loresheets.json": {"loresheets": len(loresheets["loresheets"])},
            "background_catalog.json": {"entries": len(background_catalog)},
            "generation_blood_potency.json": {
                "generations": len(generation["generations"]),
                "blood_potency_rows": len(generation["blood_potency"]),
            },
            "dyscrasias.json": {"resonances": len(dyscrasia_data["resonances"])},
            "predator_catalog.json": {"predators": len(predators)},
            "xp_matrix.json": {"rows": len(xp_matrix["rows"])},
            "equipment_qualities.json": {"qualities": len(equipment["qualities"])},
            "clans_table.json": {"clans": len(clans)},
            "discipline_powers.json": {
                "disciplines": len(discipline_data["disciplines"]),
                "powers": sum(d["power_count"] for d in discipline_data["disciplines"]),
                "quick_reference": len(discipline_data["quick_reference"]),
            },
        },
    }

    if args.dry_run:
        print(json.dumps(manifest, indent=2))
        wb.close()
        return 0

    outputs = {
        "merits_flaws.json": merits_flaws,
        "loresheets.json": loresheets,
        "background_catalog.json": background_data,
        "generation_blood_potency.json": generation_data,
        "dyscrasias.json": dyscrasia_data,
        "predator_catalog.json": predator_catalog,
        "xp_matrix.json": xp_matrix,
        "equipment_qualities.json": equipment,
        "clans_table.json": clans_table,
        "discipline_powers.json": discipline_data,
        "import_manifest.json": manifest,
    }
    for filename, payload in outputs.items():
        write_json(DATA / filename, payload)

    sync_legacy_merits(merits_flaws, source_ref)
    sync_thin_blood_merits(merits_flaws, source_ref)
    sync_thin_blood_formulas(discipline_data, source_ref)

    wb.close()
    print(f"Imported from {xlsx} (sheet {version})")
    print(json.dumps(manifest["files"], indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
